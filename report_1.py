'''
this is a variation of the code I wrote to automate formatting a report. I have removed confidential information and added some examples of what you can do with it.
'''
import openpyxl
from openpyxl.styles import Alignment

# Load the workbook
file_path = 'C:\\Users\\USER\\...\\report_1.xlsx'
wb = openpyxl.load_workbook(file_path)

# Access the first sheet in the workbook
sheet = wb.active

# Rename the sheet to what you want
sheet.title = 'report_1'

# Below are examples of manipulating the spread sheet
# Cut Column Q and insert it in front of Column A
col_q = [cell.value for cell in sheet['Q']]
sheet.insert_cols(1)  # Insert a new column at position 1 (A)
for idx, value in enumerate(col_q, start=1):
sheet.cell(row=idx, column=1, value=value)
sheet.delete_cols(18)  # Delete the original column Q (now at position 18)

# Wrap text for all cells except for the first row
for row in sheet.iter_rows(min_row=2):
for cell in row:
cell.alignment = Alignment(wrap_text=True)

# Center align columns excluding the first row
for row in sheet.iter_rows(min_row=2):
for cell in row:
if cell.column_letter in ['D', 'E', 'F', 'I', 'M', 'N', 'J']:
cell.alignment = Alignment(horizontal='center')

# Add commas for columns G and H excluding the first row
for row in sheet.iter_rows(min_row=2):
for cell in row:
if cell.column_letter in ['G', 'H', 'I']:
cell.number_format = '#,##0'

# Format columns J, K, and L as currency excluding the first row
for row in sheet.iter_rows(min_row=2):
for cell in row:
if cell.column_letter in ['K', 'L', 'M']:
cell.number_format = '"$"#,##0.00'

# Save the modified workbook
wb.save(C:\\Users\\USER\\...\\report_1.xlsx)
